"""
WSL Analytics - Missing Data Updater
=====================================
Run this script on YOUR OWN MACHINE to scrape the missing
GCA (Goal & Shot Creation) and Passing data from fbref.com
and add it to your WSL_Analytics_2025-26.xlsx file.

Requirements:
    pip install requests beautifulsoup4 pandas openpyxl lxml

Usage:
    1. Place this script in the same folder as WSL_Analytics_2025-26.xlsx
    2. Run:  python update_wsl_data.py
    3. The script will update WSL_Analytics_2025-26.xlsx in place.
"""

import os
import time
import re
import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ──────────────────────────────────────────────────────
EXCEL_FILE = "WSL_Analytics_2025-26.xlsx"

URLS = {
    "Player GCA":     "https://fbref.com/en/comps/189/gca/Womens-Super-League-Stats",
    "Player Passing": "https://fbref.com/en/comps/189/passing/Womens-Super-League-Stats",
}

# Table IDs on fbref pages
TABLE_IDS = {
    "Player GCA":     "stats_gca",
    "Player Passing": "stats_passing",
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "DNT": "1",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Sec-Fetch-User": "?1",
    "Cache-Control": "max-age=0",
}

# ── Styling helpers ──────────────────────────────────────────────
HDR_BG  = "1A237E"
ALT_ROW = "E8EAF6"

def thin_border():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font       = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill       = PatternFill("solid", fgColor=HDR_BG)
        cell.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border     = thin_border()

def style_row(ws, row, ncols, alt=False):
    bg = ALT_ROW if alt else "FFFFFF"
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font      = Font(name="Arial", size=9)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border()
    # left-align player name (col 2)
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="center")

def autowidth(ws):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value or "")) for c in col), default=6)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 6), 30)

# ── Scraper ──────────────────────────────────────────────────────
def scrape_table(url, table_id, sheet_name=None):
    """Fetch fbref page and parse the specified table into a DataFrame."""
    html_content = None
    
    # Try local file first if sheet_name is provided
    if sheet_name:
        local_file = f"{sheet_name.replace(' ', '_')}.html"
        if os.path.exists(local_file):
            print(f"  Found local file: {local_file}")
            with open(local_file, "r", encoding="utf-8") as f:
                html_content = f.read()
    
    if not html_content:
        print(f"  Fetching: {url}")
        try:
            resp = requests.get(url, headers=HEADERS, timeout=30)
            resp.raise_for_status()
            html_content = resp.text
        except requests.exceptions.HTTPError as e:
            if resp.status_code == 403:
                print(f"  ERROR: 403 Forbidden (Cloudflare block).")
                print(f"  WORKAROUND: Please save the FBref page manually as an HTML file named '{sheet_name.replace(' ', '_')}.html' in this folder and run again.")
                return None
            raise e

    soup = BeautifulSoup(html_content, "lxml")

    # fbref sometimes wraps tables in comments
    table = soup.find("table", {"id": table_id})
    if table is None:
        # Try searching inside HTML comments
        for comment in soup.find_all(string=lambda t: isinstance(t, str) and table_id in t):
            inner = BeautifulSoup(comment, "lxml")
            table = inner.find("table", {"id": table_id})
            if table:
                break

    if table is None:
        print(f"  WARNING: Could not find table '{table_id}' on {url}")
        return None

    # Parse with pandas
    df = pd.read_html(str(table), header=[0, 1])[0]

    # Flatten multi-level columns
    df.columns = [
        " ".join(str(c).strip() for c in col if "Unnamed" not in str(c)).strip()
        for col in df.columns
    ]

    # Drop repeated header rows
    df = df[df.iloc[:, 0].astype(str).str.strip() != df.columns[0]]
    df = df[df.iloc[:, 0].astype(str).str.match(r"^\d+$")]
    df.reset_index(drop=True, inplace=True)

    # Drop 'Matches' column if present
    df = df[[c for c in df.columns if "Matches" not in c]]

    print(f"  Scraped {len(df)} rows, {len(df.columns)} columns")
    return df

# ── Write to Excel ────────────────────────────────────────────────
def write_sheet(wb, sheet_name, df, tab_color="1565C0"):
    """Add or replace a sheet in the workbook with the given DataFrame."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = tab_color
    ws.freeze_panes = "C2"
    ws.row_dimensions[1].height = 32

    cols = list(df.columns)
    ncols = len(cols)

    # Header row
    for c, h in enumerate(cols, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, ncols)

    # Data rows
    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        for c_idx, val in enumerate(row, 1):
            # Convert numeric strings
            try:
                val = float(val) if "." in str(val) else int(val)
            except (ValueError, TypeError):
                pass
            ws.cell(row=r_idx, column=c_idx, value=val)
        style_row(ws, r_idx, ncols, alt=(r_idx % 2 == 1))

    autowidth(ws)
    ws.column_dimensions["B"].width = 26  # player name
    if "E" in [get_column_letter(i) for i in range(1, ncols + 1)]:
        ws.column_dimensions["E"].width = 24  # squad

    print(f"  Written sheet '{sheet_name}' ({len(df)} rows)")

# ── Main ─────────────────────────────────────────────────────────
def main():
    print(f"\nLoading workbook: {EXCEL_FILE}")
    wb = load_workbook(EXCEL_FILE)

    for sheet_name, url in URLS.items():
        table_id = TABLE_IDS[sheet_name]
        print(f"\n[{sheet_name}]")
        try:
            df = scrape_table(url, table_id, sheet_name=sheet_name)
            if df is not None and not df.empty:
                color = "006064" if "GCA" in sheet_name else "00695C"
                write_sheet(wb, sheet_name, df, tab_color=color)
            else:
                print(f"  Skipped (no data returned)")
        except Exception as e:
            print(f"  ERROR: {e}")

        time.sleep(4)  # be polite to fbref servers

    wb.save(EXCEL_FILE)
    print(f"\nDone! Saved to: {EXCEL_FILE}")
    print("Open the file and check the new 'Player GCA' and 'Player Passing' sheets.")

if __name__ == "__main__":
    main()
