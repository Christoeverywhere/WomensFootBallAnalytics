"""
Women's Football Injury Prediction Model
==========================================
Predicts injury risk for individual players or all players across 3 horizons:
  - Next match
  - Next 7 days
  - Next 30 days

Also predicts injury severity (days out proxy).

SETUP:
    pip install xgboost scikit-learn pandas numpy openpyxl

RUN (predict all players):
    python injury_prediction_model.py

RUN (predict specific player):
    python injury_prediction_model.py --player "Sam Kerr"

OUTPUT:
    injury_predictions.xlsx
"""

import argparse
import ast
import warnings
import pickle
import os
import numpy as np
import pandas as pd
warnings.filterwarnings("ignore")

from sklearn.model_selection import train_test_split, StratifiedKFold, cross_val_score
from sklearn.linear_model  import LogisticRegression
from sklearn.ensemble      import RandomForestClassifier
from sklearn.preprocessing import StandardScaler
from sklearn.metrics       import roc_auc_score, f1_score, accuracy_score
from sklearn.calibration   import CalibratedClassifierCV
from xgboost               import XGBClassifier, XGBRegressor
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils  import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
LINEUPS_PATH = "./lineups_master.csv"
MATCHES_PATH = "./matches_master.csv"
PS_PATH      = "./player_summary.csv"
MODEL_CACHE  = "./injury_model_cache.pkl"
OUTPUT_XLSX  = "./injury_predictions.xlsx"

FEATURES = [
    'minutes_played','is_starter','subbed_off',
    'yellow_cards','red_cards','position_risk',
    'days_since_last_match','high_congestion','med_congestion',
    'matches_last_7d','minutes_last_7d',
    'matches_last_14d','minutes_last_14d',
    'matches_last_30d','minutes_last_30d',
    'physical_load','aggression','pressure','duel',
    'block','clearance','interception','foul_committed',
    'miscontrol','dispossessed','pass','shot','carry',
]

HIGH_RISK_POS = ['Goalkeeper','Center Back','Right Back','Left Back',
                 'Right Center Back','Left Center Back','Center Forward',
                 'Right Wing','Left Wing','Right Center Forward','Left Center Forward']
MED_RISK_POS  = ['Center Midfield','Right Midfield','Left Midfield',
                 'Right Defensive Midfield','Left Defensive Midfield',
                 'Center Defensive Midfield','Right Center Midfield','Left Center Midfield']


# ── Helpers ───────────────────────────────────────────────────────────────────
def parse_minutes(pos_str):
    try:
        p = ast.literal_eval(str(pos_str))
        if not p: return 0
        to = p[-1].get('to')
        if to is None: return 90
        m, s = map(int, str(to).split(':'))
        return m + s/60
    except: return 0

def parse_position(pos_str):
    try:
        p = ast.literal_eval(str(pos_str))
        return p[0].get('position','Unknown') if p else 'Unknown'
    except: return 'Unknown'

def was_starter(pos_str):
    try:
        p = ast.literal_eval(str(pos_str))
        return 1 if p and p[0].get('start_reason','') == 'Starting XI' else 0
    except: return 0

def was_subbed_off(pos_str):
    try:
        p = ast.literal_eval(str(pos_str))
        return 1 if p and 'Substitution' in str(p[-1].get('end_reason','')) else 0
    except: return 0

def get_cards(card_str):
    try:
        cards = ast.literal_eval(str(card_str))
        return (sum(1 for c in cards if 'Yellow' in str(c.get('card_type',''))),
                sum(1 for c in cards if 'Red'    in str(c.get('card_type',''))))
    except: return 0, 0

def pos_risk(p):
    if p in HIGH_RISK_POS: return 2
    if p in MED_RISK_POS:  return 1
    return 0

def risk_tier(score):
    if score >= 65: return 'HIGH'
    if score >= 35: return 'MEDIUM'
    return 'LOW'


# ── Build features ────────────────────────────────────────────────────────────
def build_features():
    print("Loading data...")
    lineups = pd.read_csv(LINEUPS_PATH)
    matches = pd.read_csv(MATCHES_PATH)
    ps      = pd.read_csv(PS_PATH)

    matches['match_date'] = pd.to_datetime(matches['match_date'])

    print("Parsing positions and minutes...")
    lineups['minutes_played']   = lineups['positions'].apply(parse_minutes)
    lineups['primary_position'] = lineups['positions'].apply(parse_position)
    lineups['is_starter']       = lineups['positions'].apply(was_starter)
    lineups['subbed_off']       = lineups['positions'].apply(was_subbed_off)
    lineups['yellow_cards'], lineups['red_cards'] = zip(*lineups['cards'].apply(get_cards))
    lineups['position_risk']    = lineups['primary_position'].apply(pos_risk)

    lineups = lineups.merge(
        matches[['match_id','match_date','competition_name','season_name']],
        on='match_id', how='left'
    ).sort_values(['player_name','match_date']).reset_index(drop=True)

    lineups['days_since_last_match'] = (
        lineups.groupby('player_name')['match_date'].diff().dt.days.fillna(14)
    )
    lineups['high_congestion'] = (lineups['days_since_last_match'] < 4).astype(int)
    lineups['med_congestion']  = ((lineups['days_since_last_match'] >= 4) &
                                   (lineups['days_since_last_match'] < 7)).astype(int)

    print("Rolling workload windows...")
    def rolling(group):
        group = group.sort_values('match_date').copy()
        for col in ['matches_last_7d','minutes_last_7d','matches_last_14d',
                    'minutes_last_14d','matches_last_30d','minutes_last_30d']:
            group[col] = 0.0
        for idx in group.index:
            d = group.at[idx,'match_date']
            for days, mc, minc in [(7,'matches_last_7d','minutes_last_7d'),
                                    (14,'matches_last_14d','minutes_last_14d'),
                                    (30,'matches_last_30d','minutes_last_30d')]:
                past = group[(group['match_date'] >= d - pd.Timedelta(days=days)) &
                             (group['match_date'] < d)]
                group.at[idx, mc]   = len(past)
                group.at[idx, minc] = past['minutes_played'].sum()
        return group

    lineups = lineups.groupby('player_name', group_keys=False).apply(rolling)

    ps2 = ps.rename(columns={
        'player':'player_name','Ball Receipt*':'ball_receipts','50/50':'duels_50_50',
        'Bad Behaviour':'bad_behaviour','Foul Committed':'foul_committed',
        'Ball Recovery':'ball_recovery','Injury Stoppage':'injury_stoppages',
        'Dribbled Past':'dribbled_past',
    })
    ps2.columns = [c.lower().replace(' ','_') for c in ps2.columns]
    ps2['physical_load'] = (ps2.get('pressure',0).fillna(0) + ps2.get('duel',0).fillna(0) +
                             ps2.get('block',0).fillna(0)    + ps2.get('clearance',0).fillna(0) +
                             ps2.get('interception',0).fillna(0) + ps2.get('dribble',0).fillna(0))
    ps2['aggression']    = (ps2.get('foul_committed',0).fillna(0)*2 +
                             ps2.get('bad_behaviour',0).fillna(0)*3 +
                             ps2.get('duel',0).fillna(0))

    lineups = lineups.merge(
        ps2[['player_name','competition_name','season_name','physical_load','aggression',
             'pressure','duel','block','clearance','interception','foul_committed',
             'injury_stoppages','miscontrol','dispossessed','pass','shot','carry']],
        on=['player_name','competition_name','season_name'], how='left'
    )
    for col in ['physical_load','aggression','pressure','duel','block','clearance',
                'interception','foul_committed','injury_stoppages','miscontrol',
                'dispossessed','pass','shot','carry']:
        lineups[col] = lineups[col].fillna(0)

    # Labels
    lineups['injury_event'] = (lineups['injury_stoppages'] > 0).astype(int)

    def future_label(group, days, colname):
        group = group.sort_values('match_date').copy()
        labels = []
        for idx, row in group.iterrows():
            future = group[(group['match_date'] > row['match_date']) &
                           (group['match_date'] <= row['match_date'] + pd.Timedelta(days=days))]
            labels.append(int(future['injury_event'].sum() > 0))
        group[colname] = labels
        return group

    def next_match_label(group):
        group = group.sort_values('match_date').copy()
        group['injury_next_match'] = group['injury_event'].shift(-1).fillna(0).astype(int)
        return group

    print("Building labels...")
    lineups = lineups.groupby('player_name', group_keys=False).apply(next_match_label)
    lineups = lineups.groupby('player_name', group_keys=False).apply(lambda g: future_label(g, 7,  'injury_next_7d'))
    lineups = lineups.groupby('player_name', group_keys=False).apply(lambda g: future_label(g, 30, 'injury_next_30d'))

    return lineups


# ── Train models ──────────────────────────────────────────────────────────────
def train_models(df):
    print("\nTraining models...")
    trained = {}
    metrics = {}

    TARGETS = {
        'next_match': 'injury_next_match',
        'next_7d':    'injury_next_7d',
        'next_30d':   'injury_next_30d',
    }

    scaler = StandardScaler()
    X_all  = df[FEATURES].fillna(0).values
    scaler.fit(X_all)

    for horizon, target in TARGETS.items():
        print(f"\n  Horizon: {horizon}")
        X = df[FEATURES].fillna(0).values
        y = df[target].values
        X_tr, X_te, y_tr, y_te = train_test_split(X, y, test_size=0.2,
                                                    random_state=42, stratify=y)
        sp = (y==0).sum() / max((y==1).sum(), 1)

        models_def = {
            'lr':  LogisticRegression(max_iter=1000, random_state=42, class_weight='balanced'),
            'rf':  RandomForestClassifier(n_estimators=200, random_state=42,
                                           class_weight='balanced', n_jobs=-1),
            'xgb': XGBClassifier(n_estimators=200, random_state=42,
                                  scale_pos_weight=sp, eval_metric='logloss', verbosity=0),
        }

        horizon_models = {}
        for mname, model in models_def.items():
            X_tr_use = scaler.transform(X_tr) if mname == 'lr' else X_tr
            X_te_use = scaler.transform(X_te) if mname == 'lr' else X_te
            cal = CalibratedClassifierCV(model, cv=3, method='isotonic')
            cal.fit(X_tr_use, y_tr)
            y_prob = cal.predict_proba(X_te_use)[:,1]
            y_pred = cal.predict(X_te_use)
            auc = round(roc_auc_score(y_te, y_prob)*100, 1)
            f1  = round(f1_score(y_te, y_pred, zero_division=0)*100, 1)
            acc = round(accuracy_score(y_te, y_pred)*100, 1)
            print(f"    {mname:<5} AUC={auc}%  F1={f1}%  Acc={acc}%")
            horizon_models[mname] = cal
            if mname not in metrics:
                metrics[mname] = {}
            metrics[mname][horizon] = {'auc': auc, 'f1': f1, 'acc': acc}

        trained[horizon] = horizon_models

    # Severity model
    sev_df = df[df['injury_stoppages'] > 0]
    sev_model = None
    if len(sev_df) > 100:
        sev_model = XGBRegressor(n_estimators=100, random_state=42, verbosity=0)
        sev_model.fit(sev_df[FEATURES].fillna(0), sev_df['injury_stoppages'])
        print("\n  Severity model trained")

    # Feature importance
    xgb_imp = XGBClassifier(n_estimators=200, random_state=42, eval_metric='logloss', verbosity=0)
    xgb_imp.fit(df[FEATURES].fillna(0), df['injury_next_match'])
    importances = pd.Series(xgb_imp.feature_importances_, index=FEATURES).sort_values(ascending=False)

    return trained, scaler, metrics, sev_model, importances


# ── Predict ───────────────────────────────────────────────────────────────────
def predict(df, trained_models, scaler, sev_model, player_name=None):
    if player_name:
        df = df[df['player_name'].str.lower() == player_name.lower()]
        if len(df) == 0:
            print(f"Player '{player_name}' not found.")
            return pd.DataFrame()

    X     = df[FEATURES].fillna(0).values
    X_sc  = scaler.transform(X)
    preds = df[['player_name','country','team_name','primary_position',
                'match_date','competition_name','season_name',
                'minutes_played','days_since_last_match',
                'matches_last_7d','matches_last_30d',
                'physical_load','aggression','high_congestion']].copy()

    for horizon, models in trained_models.items():
        lr_p  = models['lr'].predict_proba(X_sc)[:,1]
        rf_p  = models['rf'].predict_proba(X)[:,1]
        xgb_p = models['xgb'].predict_proba(X)[:,1]
        ens_p = np.mean([lr_p, rf_p, xgb_p], axis=0)

        preds[f'{horizon}_injury_prob_pct'] = (ens_p * 100).round(1)
        preds[f'{horizon}_risk_tier']       = [risk_tier(s) for s in ens_p * 100]
        preds[f'{horizon}_lr_prob']         = (lr_p  * 100).round(1)
        preds[f'{horizon}_rf_prob']         = (rf_p  * 100).round(1)
        preds[f'{horizon}_xgb_prob']        = (xgb_p * 100).round(1)

    if sev_model:
        preds['severity_score'] = sev_model.predict(X).round(1)
    else:
        preds['severity_score'] = 0

    preds['match_date'] = pd.to_datetime(preds['match_date']).dt.strftime('%Y-%m-%d')
    return preds


# ── Excel output ──────────────────────────────────────────────────────────────
def save_excel(all_preds, latest_preds, metrics, importances, output_path):
    print(f"\nSaving Excel report → {output_path}")
    NAVY  = PatternFill("solid", fgColor="1A3A5C")
    RED   = PatternFill("solid", fgColor="C8252A")
    BLUE  = PatternFill("solid", fgColor="2E6DA4")
    ALT   = PatternFill("solid", fgColor="EBF2FA")
    WHITE = PatternFill("solid", fgColor="FFFFFF")
    H_RED = PatternFill("solid", fgColor="FFD7D7")
    H_ORG = PatternFill("solid", fgColor="FFF0CD")
    H_GRN = PatternFill("solid", fgColor="D4EDDA")
    thin  = Side(style='thin', color='CCCCCC')
    BORD  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hf(s=11, bold=True): return Font(name='Calibri', bold=bold, color="FFFFFF", size=s)
    def bf(s=10): return Font(name='Calibri', size=s)

    def write_sheet(wb, name, df, title, subtitle):
        ws = wb.create_sheet(title=name)
        ws.sheet_view.showGridLines = False
        n = len(df.columns)
        ws.merge_cells(f"A1:{get_column_letter(n)}1")
        ws['A1'] = title; ws['A1'].font = hf(14)
        ws['A1'].fill = RED
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 32
        ws.merge_cells(f"A2:{get_column_letter(n)}2")
        ws['A2'] = subtitle; ws['A2'].font = hf(10, False)
        ws['A2'].fill = BLUE
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 18
        for ci, col in enumerate(df.columns, 1):
            c = ws.cell(row=3, column=ci, value=col.replace('_',' ').title())
            c.font = hf(); c.fill = NAVY
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = BORD
        ws.row_dimensions[3].height = 22
        for ri, row in enumerate(df.itertuples(index=False), 4):
            fill = ALT if ri % 2 == 0 else WHITE
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = bf(); c.border = BORD
                c.alignment = Alignment(horizontal='center' if ci > 2 else 'left', vertical='center')
                col_name = df.columns[ci-1]
                if 'risk_tier' in col_name:
                    c.fill = H_RED if val=='HIGH' else H_ORG if val=='MEDIUM' else H_GRN if val=='LOW' else fill
                elif 'prob_pct' in col_name or 'prob' in col_name:
                    try:
                        v = float(val)
                        c.fill = H_RED if v>=65 else H_ORG if v>=35 else H_GRN
                    except: c.fill = fill
                else: c.fill = fill
            ws.row_dimensions[ri].height = 15
        for ci, col in enumerate(df.columns, 1):
            w = max(len(str(col)), df[col].astype(str).str.len().max() if len(df)>0 else 8)
            ws.column_dimensions[get_column_letter(ci)].width = min(max(w+2,10),28)
        return ws

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: All players latest
    cols = ['player_name','country','team_name','primary_position',
            'next_match_injury_prob_pct','next_match_risk_tier',
            'next_7d_injury_prob_pct','next_7d_risk_tier',
            'next_30d_injury_prob_pct','next_30d_risk_tier',
            'severity_score','minutes_played','matches_last_30d',
            'physical_load','high_congestion','days_since_last_match']
    write_sheet(wb, "🏥 All Player Predictions",
                latest_preds[cols].sort_values('next_match_injury_prob_pct', ascending=False).reset_index(drop=True),
                "All Players — Injury Risk Predictions",
                "Sorted by next-match injury probability | Ensemble model (LR + RF + XGBoost)")

    # Sheet 2: High risk
    high = latest_preds[latest_preds['next_match_risk_tier']=='HIGH'][cols].sort_values(
        'next_match_injury_prob_pct', ascending=False).reset_index(drop=True)
    write_sheet(wb, "🔴 High Risk",  high,  "HIGH RISK Players",  f"{len(high)} players — 65%+ injury probability")

    # Sheet 3: Medium risk
    med = latest_preds[latest_preds['next_match_risk_tier']=='MEDIUM'][cols].sort_values(
        'next_match_injury_prob_pct', ascending=False).reset_index(drop=True)
    write_sheet(wb, "🟡 Medium Risk", med, "MEDIUM RISK Players", f"{len(med)} players — 35–64% injury probability")

    # Sheet 4: Per-match top 5000
    pm_cols = ['player_name','team_name','primary_position','match_date','competition_name',
               'minutes_played','days_since_last_match','high_congestion','physical_load',
               'next_match_injury_prob_pct','next_match_risk_tier',
               'next_7d_injury_prob_pct','next_30d_injury_prob_pct','severity_score']
    write_sheet(wb, "📋 Per-Match (Top 5000)",
                all_preds[pm_cols].sort_values('next_match_injury_prob_pct', ascending=False).head(5000).reset_index(drop=True),
                "Per-Match Predictions — Top 5,000 Highest Risk",
                "Full dataset: 21,172 rows across all matches")

    # Sheet 5: Feature importance
    imp_df = importances.reset_index()
    imp_df.columns = ['Feature','Score']
    imp_df['Pct'] = (imp_df['Score'] / imp_df['Score'].sum() * 100).round(1)
    imp_df.insert(0,'Rank', range(1, len(imp_df)+1))
    write_sheet(wb, "🔍 Feature Importance", imp_df,
                "Feature Importance (XGBoost)", "Which features matter most for injury prediction")

    wb.save(output_path)
    print(f"  ✅ Saved: {output_path}")


# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Women's Football Injury Prediction")
    parser.add_argument('--player', type=str, default=None, help='Predict for specific player name')
    parser.add_argument('--retrain', action='store_true', help='Force retrain even if cache exists')
    parser.add_argument('--list',   action='store_true', help='List available players (use with --player to search)')
    args = parser.parse_args()

    # Build features
    df = build_features()

    # Train or load from cache
    if os.path.exists(MODEL_CACHE) and not args.retrain:
        print(f"\nLoading cached model from {MODEL_CACHE}")
        with open(MODEL_CACHE, 'rb') as f:
            cache = pickle.load(f)
        trained, scaler, metrics, sev_model, importances = (
            cache['trained'], cache['scaler'], cache['metrics'],
            cache['sev_model'], cache['importances'])
    else:
        trained, scaler, metrics, sev_model, importances = train_models(df)
        with open(MODEL_CACHE, 'wb') as f:
            pickle.dump({'trained': trained, 'scaler': scaler, 'metrics': metrics,
                         'sev_model': sev_model, 'importances': importances}, f)
        print(f"\nModel cached → {MODEL_CACHE}")

    # ── List mode ─────────────────────────────────────────────────────────────
    if args.list:
        all_names = sorted(df['player_name'].unique())
        if args.player:
            # Filter by search term
            matches_names = [n for n in all_names if args.player.lower() in n.lower()]
            print(f"\nPlayers matching '{args.player}':")
            for n in matches_names:
                print(f"  {n}")
        else:
            print(f"\n{len(all_names)} players in dataset (first 50):")
            for n in all_names[:50]:
                print(f"  {n}")
            print("  ...use --list --player <name> to search")
        exit(0)

    # ── Predict ────────────────────────────────────────────────────────────────
    print("\nGenerating predictions...")
    all_preds = predict(df, trained, scaler, sev_model, player_name=args.player)

    if args.player:
        if len(all_preds) == 0:
            # Fuzzy suggest
            all_names = sorted(df['player_name'].unique())
            suggestions = [n for n in all_names if
                           any(part.lower() in n.lower()
                               for part in args.player.split())][:5]
            print(f"\n❌ Player '{args.player}' not found in dataset.")
            print("   This dataset covers WSL, NWSL, Women's World Cup and UEFA Women's Euro only.")
            if suggestions:
                print(f"\n   Did you mean one of these?")
                for s in suggestions:
                    print(f"     → {s}")
            print("\n   Use:  python injury_prediction_model.py --list --player <lastname>")
            print("         to search for a player by name.")
            exit(1)

        # Print summary to terminal
        latest = all_preds.sort_values('match_date').tail(1)
        print(f"\n{'='*55}")
        print(f"  Predictions for: {args.player}")
        print(f"{'='*55}")
        for horizon in ['next_match','next_7d','next_30d']:
            prob = latest[f'{horizon}_injury_prob_pct'].values[0]
            tier = latest[f'{horizon}_risk_tier'].values[0]
            print(f"  {horizon:<12}: {prob}%  [{tier}]")
        print(f"  Severity score : {latest['severity_score'].values[0]}")
        print(f"{'='*55}")
        output_path = f"./injury_pred_{args.player.replace(' ','_')}.xlsx"
    else:
        output_path = OUTPUT_XLSX

    latest_preds = all_preds.sort_values('match_date').groupby('player_name').last().reset_index()
    save_excel(all_preds, latest_preds, metrics, importances, output_path)
    print(f"\n✅ Done! Open: {output_path}")